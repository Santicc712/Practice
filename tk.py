from tkinter import *
from tkinter import ttk
from tkinter.messagebox import showinfo
from openpyxl import Workbook, load_workbook
import os

root = Tk()  # создаем корневой объект - окно
root.title("Запись клиентов")  # устанавливаем заголовок окна
root.geometry("560x560+600+200")  # устанавливаем размеры окна


# Функция для записи данных в Excel файл
def accept():
    name = entry_name.get()
    date = entry_date.get()
    description = editor.get("1.0", END).strip()

    # Проверяем, чтобы поля не были пустыми
    if not name or not date or not description:
        showinfo(title='Ошибка', message='Все поля должны быть заполнены!')
        return

    file_name = 'client_records.xlsx'

    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Имя", "Дата", "Описание"])  # добавляем заголовки столбцов при создании нового файла

    sheet.append([name, date, description])

    # Автоматически устанавливаем ширину столбцов по содержимому
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # получаем букву столбца
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(file_name)
    workbook.close()

    showinfo(title='Информация', message='Запись создана')
    entry_name.delete(0, END)
    entry_date.delete(0, END)
    editor.delete("1.0", END)


# Функция для открытия файла с записями
def open_journal():
    file_name = 'client_records.xlsx'
    if os.path.exists(file_name):
        os.startfile(file_name)
    else:
        showinfo(title='Информация', message='Файл с записями не найден.')


Name = Label(text="Имя:", font=("Arial", 14))  # создаем текстовую метку
Name.pack(anchor=W, ipady=10)  # размещаем метку в окне

entry_name = ttk.Entry(font=("Arial", 15))
entry_name.pack(anchor=W)

Date = Label(text="Дата:", font=("Arial", 14))
Date.pack(anchor=W, ipady=10)

entry_date = ttk.Entry(font=("Arial", 15))
entry_date.pack(anchor=W)

Work = Label(text="Описание:", font=("Arial", 14))
Work.pack(anchor=W, ipady=10)

btn = ttk.Button(text='Записать', command=accept)
btn1 = ttk.Button(text='Журнал записей', command=open_journal)
btn1.pack(side=BOTTOM, fill=BOTH)
btn.pack(side=BOTTOM, fill=BOTH)  # размещаем кнопку в окне

editor = Text(font=("Arial", 15))
editor.pack(fill=BOTH, expand=1)

root.mainloop()
